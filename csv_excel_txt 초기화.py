#컴파일 하지 마라 ....
import csv
import openpyxl
import os

#requirement 까는법
# 0) python -m venv env -> activate한 다음
# 1) pip install -r requirement.txt

folderpath = ""

#폴더내 파일이 없다면
if not os.path.exists(folderpath):
  pass
else:
  pass #이 밑에부터 써야 함

#csv+txt========================================
# 0)만들기
f = open(folderpath,'w',encoding='cp949 or utf-8', newline='')
writer = csv.writer(f)

# 1)불러오기
f = open(folderpath,'r',encoding='cp949 or utf-8')
reader = csv.reader(f)

# 끝내기
f.close()#파라미터 안들어감

#엑셀==============================================
#자세한건 z010-삼전+엑셀.py 참조
wb = openpyxl.Workbook()#만들기 주의!! 파라미터 없음
wb = openpyxl.load_workbook(folderpath)#불러오기

# 엑셀 시트 만들기
ws = wb.active#활성화된 시트 
ws = wb.create_sheet('종가')#시트만들기
ws["A1"] = ""


wb.save(folderpath)#파라미터 들어감