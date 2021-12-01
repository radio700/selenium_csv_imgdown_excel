import requests as req
from bs4 import BeautifulSoup as bs
import openpyxl
import time
import os
from datetime import datetime

#https://www.youtube.com/watch?v=n8_SeIZ_vo4&t=5s

path = r'C:\Users\jaws\Desktop\crwal\project\pure_exersice\금감원자료.xlsx'

if not os.path.exists(path):
    print('워크북을 만듭니다')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "종가"
    ws['A1'] = '날짜'
    ws['B1'] = '참가번호'
    ws['C1'] = '현재가'
    ws['D1'] = '평균매입가'
    ws['E1'] = "잔고수량"
    ws['F1'] = "평가금액"
    ws['G1'] = "평가손익"
    ws['h1'] = "수익률"
    wb.save(path)
else:
    print('이제부터 씁니다')
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    codes = {
        '005930': '삼성전자',
        '000660': 'sk하이닉스',
        '035720': 'kakao',
    }

    row = 2


    for code in codes:

        url = f"https://finance.naver.com/item/sise.naver?code={code}"

        res = req.get(url)
        html = res.text
        soup = bs(html, 'html.parser')

        type = codes[f'{code}']
        print(f'{type}입니다')

        price = soup.select_one('#_nowVal').text
        price = int(price.replace(',', ''))
        print(price)

        year = datetime.now().year
        month = datetime.now().month
        day = datetime.now().day

        ws[f'A{row}'] = (f"{year}/{month}/{day}")

        ws[f'B{row}'] = type

        ws[f"C{row}"] = int(price)
        row = row+1

    wb.save(path)

    # select는 리스트로 반환돼서 뒤에 .text가 붙지 않는다
    # 결과값 : [<strong class="tah p11" id="_nowVal">71,200</strong>] 이 나오고
    # price = soup.select('#_nowVal')
    # print(price)

    # select_one은 리스트로 반환되지 않고 객체로 떼온다
    # 결과값 : <strong class="tah p11" id="_nowVal">71,200</strong>
    # price = soup.select_one('#_nowVal')
    # print(price)