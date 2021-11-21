import openpyxl

#https://www.youtube.com/watch?v=n8_SeIZ_vo4&t=5s

# 불러오기
path = r'C:\Users\jaws\Desktop\crwal\project\pure_exersice\eeee.xlsx'

wb = openpyxl.load_workbook(path)

# 엑셀 시트선택
ws = wb['종가']

# 데이터 수정
ws['A3'] = 456
ws['B3'] = '성기훈'

# 데이터 저장

wb.save(path)