import openpyxl

#https://www.youtube.com/watch?v=n8_SeIZ_vo4&t=5s

# 1 엑셀 만드릭
wb = openpyxl.Workbook()

# 엑셀 시트 만들기
ws = wb.create_sheet('종가')

# 데이터추가
ws['A1'] = '참가번호'
ws['B1'] = '성명'

ws['A2'] = '1'
ws['B2'] = '오일날'

# 엑셀저장하기
wb.save(r'C:\Users\jaws\Desktop\crwal\project\pure_exersice\eeee.xlsx')